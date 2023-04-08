from openpyxl import load_workbook


class ExcelParser:
    """Class for parsing Excel files"""
    def __init__(self, filename):
        self.filename = filename

    def parse_worksheet(self):
        workbook = load_workbook(filename=self.filename)
        worksheet = workbook.active
        data_rows = worksheet.iter_rows(min_row=4, values_only=True)
        return data_rows
    