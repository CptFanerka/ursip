"""Use 'python excel_parser.py' command to run this script"""
import sqlite3
from datetime import datetime
import random
from openpyxl import load_workbook
from prettytable import PrettyTable


class ExcelParser:
    """Class for parsing Excel files"""
    def __init__(self, filename):
        self.filename = filename

    def parse_worksheet(self):
        workbook = load_workbook(filename=self.filename)
        worksheet = workbook.active
        data_rows = worksheet.iter_rows(min_row=4, values_only=True)
        return data_rows

class Database:
    """Class for managing SQLIte database"""
    def __init__(self, db_name):
        self.db_name = db_name
        self.connection = None
        self.cursor = None

    def __enter__(self):
        self.connection = sqlite3.connect(self.db_name)
        self.cursor = self.connection.cursor()
        return self

    def close(self):
        self.cursor.close()
        self.connection.close()

    def create_table(self):
        create_table_query = '''CREATE TABLE IF NOT EXISTS excel_data (
                                id INTEGER PRIMARY KEY,
                                company TEXT NOT NULL,
                                factQliqdata1 INTEGER NOT NULL,
                                factQliqdata2 INTEGER NOT NULL,
                                factQoildata1 INTEGER NOT NULL,
                                factQoildata2 INTEGER NOT NULL,
                                forecastQliqdata1 INTEGER NOT NULL,
                                forecastQliqdata2 INTEGER NOT NULL,
                                forecastQoildata1 INTEGER NOT NULL,
                                forecastQoildata2 INTEGER NOT NULL,
                                date INTEGER NOT NULL);'''
        self.cursor.execute(create_table_query)
        self.connection.commit()

    def insert_data(self, values):
        insert_query = f"""INSERT INTO excel_data {COLUMNS_NAMES_STR} VALUES {values}"""
        try:
            self.cursor.execute(insert_query)
        except sqlite3.IntegrityError:
            # this data already in table, ignore it
            pass
        self.connection.commit()

    def select_total_data(self):
        select_total_query = '''
                            SELECT date, company, value_type,
                                SUM(factQliqdata1 + factQliqdata2) AS TotalQliq,
                                SUM(factQoildata1 + factQoildata2) AS TotalQoil
                            FROM (
                                SELECT date, company, 'fact' AS value_type,
                                    factQliqdata1, factQliqdata2, factQoildata1, factQoildata2
                                FROM excel_data
                                UNION ALL
                                SELECT date, company, 'forecast' AS value_type,
                                    forecastQliqdata1, forecastQliqdata2, forecastQoildata1, forecastQoildata2
                                FROM excel_data
                            )
                            GROUP BY date, company, value_type
                            '''
        self.cursor.execute(select_total_query)
        results = self.cursor.fetchall()
        return results

class DataProcessor:
    """Class for appending date to Excel data"""
    def __init__(self, data_rows):
        self.data_rows = data_rows

    def process_data(self, storage):
        for row in self.data_rows:
            date = datetime(YEAR, MONTH, random.randint(1, 30))
            timestamp = int(date.timestamp())
            data = row[0:10] + (timestamp,)
            storage.insert_data(data)

class TableFormatter:
    """Class for pretty printing of result table"""
    def __init__(self, headers):
        self.table = PrettyTable()
        self.table.field_names = headers

    def add_row(self, row):
        self.table.add_row(row)

    def show(self):
        print(self.table)


if __name__ == '__main__':
    COLUMNS_NAMES_STR = ("(id, "
                    "company, "
                    "factQliqdata1, "
                    "factQliqdata2, "
                    "factQoildata1, "
                    "factQoildata2, "
                    "forecastQliqdata1, "
                    "forecastQliqdata2, "
                    "forecastQoildata1, "
                    "forecastQoildata2, "
                    "date)")
    YEAR = 2023
    MONTH = 4

    file_parser = ExcelParser(filename='file.xlsx')
    data_rows = file_parser.parse_worksheet()

    sqlite_db = Database('sqlite_python.db')
    sqlite_db.__enter__()
    sqlite_db.create_table()

    data_processor = DataProcessor(data_rows)
    data_processor.process_data(storage=sqlite_db)

    results = sqlite_db.select_total_data()

    table_headers = [description[0] for description in sqlite_db.cursor.description]
    table_formatter = TableFormatter(table_headers)
    for row in results:
        # row structure example: (1682629200, 'company1', 'forecast', 88, 180)
        date = datetime.fromtimestamp(row[0]).strftime('%Y-%m-%d')
        table_formatter.add_row([date, row[1], row[2], row[3], row[4]])
    table_formatter.show()

    sqlite_db.close()
